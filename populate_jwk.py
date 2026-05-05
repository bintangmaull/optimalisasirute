import openpyxl
from openpyxl.utils import get_column_letter

file_path = r"D:\Ayak\Project Rolling\ZRS86 MEI.XLSX"
print(f"Loading {file_path}...")

wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Get headers
headers = [cell.value for cell in ws[1]]

# Map day columns. We know TueFrom2 might be missing/None but it's in the sequence.
# Sequence in headers: MonFrom1(23), MonFrom2(24), TueFrom1(25), None(26), WedFrom1(27)...
days_config = [
    ("SENIN", "MonFrom1", "MonFrom2"),
    ("SELASA", "TueFrom1", "TueFrom2"), # Note: TueFrom2 might be None in headers but we check index 26
    ("RABU", "WedFrom1", "WedFrom2"),
    ("KAMIS", "ThuFrom1", "ThuFrom2"),
    ("JUMAT", "FriFrom1", "FriFrom2"),
    ("SABTU", "SatFrom1", "SatFrom2"),
]

# Find indices. MonFrom1 is index 23 (1-indexed: 24)
mon_from_1_idx = 0
for i, h in enumerate(headers):
    if h == "MonFrom1":
        mon_from_1_idx = i + 1
        break

if not mon_from_1_idx:
    print("Error: MonFrom1 column not found.")
    exit(1)

# Insert JWK SALESMAN before MonFrom1
ws.insert_cols(mon_from_1_idx)
ws.cell(row=1, column=mon_from_1_idx).value = "JWK SALESMAN"

# New column is at mon_from_1_idx. 
# Shift the others: MonFrom1 is now at mon_from_1_idx + 1, etc.
def get_val(row, col_name_or_idx, offset=1):
    if isinstance(col_name_or_idx, int):
        val = ws.cell(row=row, column=col_name_or_idx).value
    else:
        # Find column index dynamically after shift
        idx = -1
        for i, h in enumerate(headers):
            if h == col_name_or_idx:
                idx = i + 1
                break
        if idx >= mon_from_1_idx:
            idx += offset
        val = ws.cell(row=row, column=idx).value if idx != -1 else None
    return str(val).strip() if val is not None else ""

target_time = "08:00:00"

for row_idx in range(2, ws.max_row + 1):
    jwk_results = []
    
    # We'll use fixed offsets from mon_from_1_idx + 1
    # MonFrom1=idx+1, MonFrom2=idx+2, TueFrom1=idx+3, TueFrom2=idx+4...
    base_idx = mon_from_1_idx + 1
    
    for i, (day_name, f1_name, f2_name) in enumerate(days_config):
        idx1 = base_idx + (i * 2)
        idx2 = base_idx + (i * 2) + 1
        
        v1 = ws.cell(row=row_idx, column=idx1).value
        v2 = ws.cell(row=row_idx, column=idx2).value
        
        v1_s = str(v1).strip() if v1 else ""
        v2_s = str(v2).strip() if v2 else ""
        
        has_v1 = target_time in v1_s
        has_v2 = target_time in v2_s
        
        if has_v1 and has_v2:
            jwk_results.append(day_name)
        elif has_v1:
            jwk_results.append(f"{day_name} GANJIL")
        elif has_v2:
            jwk_results.append(f"{day_name} GENAP")
            
    ws.cell(row=row_idx, column=mon_from_1_idx).value = ", ".join(jwk_results)

wb.save(file_path)
print("Saved successfully.")
