import openpyxl
import os

file_path = r"D:\Ayak\Project Rolling\ZR268 (1).xlsx"
print(f"Loading {file_path}...")

wb = openpyxl.load_workbook(file_path)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"Unhiding columns in sheet: {sheet_name}")
    
    # Iterate through all columns that have defined dimensions
    for col_idx in range(1, ws.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].hidden = False

wb.save(file_path)
print("Saved successfully.")
