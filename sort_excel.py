import openpyxl

file_path = r"D:\Ayak\Project Rolling\ZRS86 MEI.XLSX"
print(f"Loading {file_path}...")

wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Get headers to find JWK SALESMAN index
headers = [cell.value for cell in ws[1]]
jwk_col_idx = -1
for i, h in enumerate(headers):
    if h == "JWK SALESMAN":
        jwk_col_idx = i
        break

if jwk_col_idx == -1:
    print("Error: JWK SALESMAN column not found.")
    exit(1)

# Sorting priority
priority_list = [
    "SENIN", "SENIN GANJIL", "SENIN GENAP",
    "SELASA", "SELASA GANJIL", "SELASA GENAP",
    "RABU", "RABU GANJIL", "RABU GENAP",
    "KAMIS", "KAMIS GANJIL", "KAMIS GENAP",
    "JUMAT", "JUMAT GANJIL", "JUMAT GENAP",
    "SABTU", "SABTU GANJIL", "SABTU GENAP"
]
priority_map = {val: i for i, val in enumerate(priority_list)}

def get_sort_key(row_data):
    val = str(row_data[jwk_col_idx]).strip() if row_data[jwk_col_idx] is not None else ""
    # Use 999 for items not in priority list (like empty or complex combinations)
    # Then sub-sort by the string itself
    return priority_map.get(val, 999), val

# Extract data rows
data_rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    data_rows.append(row)

# Sort data rows
data_rows.sort(key=get_sort_key)

# Delete existing data rows in sheet
# Note: delete_rows is more efficient than clearing cells one by one
if ws.max_row > 1:
    ws.delete_rows(2, amount=ws.max_row - 1)

# Write back sorted data
for row_idx, row_data in enumerate(data_rows, start=2):
    for col_idx, cell_value in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx).value = cell_value

# Remove AutoFilter
ws.auto_filter.ref = None

wb.save(file_path)
print("Sorted and filters removed successfully.")
