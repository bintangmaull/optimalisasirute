import pandas as pd
import openpyxl

CSV_PATH = r"D:\Ayak\Project Rolling\ZRS86_Filtered.csv"
XLSX_PATH = r"D:\Ayak\Project Rolling\ZRS86 MEI.XLSX"

def finalize():
    print("Reading optimized CSV...")
    df = pd.read_csv(CSV_PATH)
    # Map (Personnel Number, Customer) -> Optimized JWK
    mapping = df.set_index(['Personnel Number', 'Customer'])['Optimized JWK'].to_dict()
    
    print(f"Loading {XLSX_PATH}...")
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    pers_col = headers.index('Personnel Number') + 1
    cust_col = headers.index('Customer') + 1
    
    # Check/Add Optimized JWK column
    if 'Optimized JWK' not in headers:
        opt_jwk_idx = (headers.index('JWK SALESMAN') + 2) if 'JWK SALESMAN' in headers else (len(headers) + 1)
        ws.insert_cols(opt_jwk_idx)
        ws.cell(row=1, column=opt_jwk_idx).value = 'Optimized JWK'
        headers.insert(opt_jwk_idx - 1, 'Optimized JWK')
    else:
        opt_jwk_idx = headers.index('Optimized JWK') + 1
        
    # Map day columns
    days_cols = {
        "SENIN": ("MonFrom1", "MonFrom2"),
        "SELASA": ("TueFrom1", "TueFrom2"),
        "RABU": ("WedFrom1", "WedFrom2"),
        "KAMIS": ("ThuFrom1", "ThuFrom2"),
        "JUMAT": ("FriFrom1", "FriFrom2"),
        "SABTU": ("SatFrom1", "SatFrom2"),
    }
    
    col_map = {}
    for d, (f1, f2) in days_cols.items():
        c1 = headers.index(f1) + 1 if f1 in headers else None
        c2 = headers.index(f2) + 1 if f2 in headers else None
        col_map[d] = (c1, c2)

    print("Updating rows...")
    TARGET_TIME = "08:00:00"
    
    for row_idx in range(2, ws.max_row + 1):
        # Personnel Number and Customer might be read as int or str from Excel
        p_num = ws.cell(row=row_idx, column=pers_col).value
        cust = ws.cell(row=row_idx, column=cust_col).value
        
        # In CSV, pandas often reads them as int64. Let's ensure types match.
        # Most robust way is to convert to int if they look like numbers.
        try:
            key = (int(p_num), int(cust))
        except:
            key = (str(p_num), str(cust))
            
        if key in mapping:
            opt_jwk = mapping[key]
            ws.cell(row=row_idx, column=opt_jwk_idx).value = opt_jwk
            
            # Clear all visit columns first
            for c1, c2 in col_map.values():
                if c1: ws.cell(row=row_idx, column=c1).value = None
                if c2: ws.cell(row=row_idx, column=c2).value = None
            
            # Set columns based on opt_jwk
            for d, (c1, c2) in col_map.items():
                if d in str(opt_jwk).upper():
                    if "GANJIL" in str(opt_jwk).upper():
                        if c1: ws.cell(row=row_idx, column=c1).value = TARGET_TIME
                    elif "GENAP" in str(opt_jwk).upper():
                        if c2: ws.cell(row=row_idx, column=c2).value = TARGET_TIME
                    else: # Weekly
                        if c1: ws.cell(row=row_idx, column=c1).value = TARGET_TIME
                        if c2: ws.cell(row=row_idx, column=c2).value = TARGET_TIME
                        
    wb.save(XLSX_PATH)
    print("Finalized Excel successfully.")

if __name__ == "__main__":
    finalize()
