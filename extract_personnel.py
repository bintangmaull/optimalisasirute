import openpyxl
import csv
import os

file_path = r"D:\Ayak\Project Rolling\ZRS86 MEI.XLSX"
output_path = r"D:\Ayak\Project Rolling\ZRS86_Filtered.csv"
target_ids = ["90051274", "90052003", "90052197", "90052353"]

print(f"Loading {file_path}...")
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb.active

headers = [cell.value for cell in ws[1]]
p_idx = -1
for i, h in enumerate(headers):
    if h == "Personnel Number":
        p_idx = i
        break

if p_idx == -1:
    print("Error: 'Personnel Number' column not found.")
    exit(1)

filtered_rows = [headers]
count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    p_num = str(row[p_idx]).strip() if row[p_idx] is not None else ""
    if p_num in target_ids:
        filtered_rows.append([str(c) if c is not None else "" for c in row])
        count += 1

with open(output_path, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerows(filtered_rows)

print(f"Extraction complete. Saved {count} rows to {output_path}")
